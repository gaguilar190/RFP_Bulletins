from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook

try:
    from rapidfuzz import fuzz, process
except Exception:
    fuzz = None
    process = None


# -------------------------------------------------------------------
# Standard field aliases
# -------------------------------------------------------------------

COLUMN_ALIASES: dict[str, list[str]] = {
    "hold_status": [
        "hold avail status",
        "hold / avail status",
        "hold status",
        "avail status",
        "availability status",
        "status",
    ],
    "market": [
        "market",
        "dma",
        "client market name",
        "client market",
        "market name",
    ],
    "vendor": [
        "vendor",
        "media owner",
        "owner",
        "supplier",
        "vendor name",
    ],
    "media_type": [
        "format",
        "media type",
        "unit type",
        "product",
        "inventory type",
        "type",
    ],
    "unit_quantity": [
        "unit quantity",
        "unit qty",
        "quantity",
        "# of units",
        "number of units",
        "of units",
    ],
    "unit_id": [
        "unit",
        "unit #",
        "unit number",
        "vendor inventory #",
        "vendor inventory number",
        "vendor inventory",
        "inventory id",
        "panel id",
        "tab id",
    ],
    "geopath_id": [
        "geopath id",
        "geopath id #",
        "geopath id number",
        "geopath frame id",
        "frame id",
    ],
    "location_description": [
        "location description",
        "location",
        "description",
        "location details",
        "location desc",
    ],
    "face": [
        "face",
        "facing",
        "read",
        "orientation",
    ],
    "latitude": [
        "latitude",
        "lat",
    ],
    "longitude": [
        "longitude",
        "long",
        "lon",
        "lng",
    ],
    "zip_code": [
        "zip",
        "zip code",
        "zipcode",
        "postal code",
    ],
    "size": [
        "size",
        "size hxw",
        "size h x w",
        "dimensions",
        "unit size",
    ],
    "digital_spot_length": [
        "digital spot length",
        "spot length",
        "spot length seconds",
    ],
    "spots_per_loop": [
        "# of spots per loop",
        "number of spots per loop",
        "spots per loop",
        "number of spots",
        "# of spots",
        "spots",
    ],
    "loop_length": [
        "loop length",
        "loop length seconds",
    ],
    "digital_display_type": [
        "digital display type",
        "display type",
    ],
    "pixel_size": [
        "pixel size",
        "pixel size h x w",
        "pixel size hxw",
        "resolution",
    ],
    "weekly_impressions": [
        "1 week a18+ geopath impressions",
        "1 week a18 geopath impressions",
        "a18+ impressions/week",
        "18+ impressions/week",
        "18 plus impressions week",
        "weekly impressions",
        "a18 weekly impressions",
        "popfacts persons 18 plus yrs 1wk total impressions",
    ],
    "four_week_impressions": [
        "4 week a18+ geopath impressions",
        "4 week a18 geopath impressions",
        "a18+ 4-wk impressions",
        "a18+ 4 wk impressions",
        "18+ impressions/cycle",
        "18 plus impressions cycle",
        "18 plus 4 week impressions",
        "4 week impressions",
    ],
    "total_impressions": [
        "18+ total impressions",
        "18 plus total impressions",
        "total impressions",
        "a18 total impressions",
    ],
    "start_date": [
        "start date",
        "start",
        "flight start",
        "posting start",
    ],
    "end_date": [
        "end date",
        "end",
        "flight end",
        "posting end",
    ],
    "cycle_duration": [
        "cycle duration",
        "cycle type",
        "period type",
    ],
    "number_of_cycles": [
        "number of cycles",
        "# of cycles",
        "cycles",
        "number of 4 wk periods",
        "# of 4 wk periods",
        "number of 4 week periods",
        "# of 4 week periods",
    ],
    "rate_card": [
        "rate card",
        "4 week rate card",
        "4wk rate card net cost",
        "4 week rate card net cost",
        "rate card/cycle",
        "rate card cycle",
    ],
    "four_week_rate": [
        "4 week rate",
        "4wk proposed net cost",
        "4 week proposed net cost",
        "4wk negotiated net cost",
        "4 week negotiated net cost",
        "4 week media cost",
        "4 week net media cost",
        "net media cost/cycle",
        "net media cost cycle",
        "net media cost",
        "proposed net cost",
        "negotiated net cost",
    ],
    "total_media_cost": [
        "total media net cost",
        "net campaign media cost",
        "total spaces cost net",
        "total media cost",
        "media total",
    ],
    "install_cost": [
        "installation cost",
        "install cost",
        "initial installation cost",
        "initial installation cost net",
        "initial install cost",
        "installation cost final",
    ],
    "production_cost": [
        "production cost",
        "production cost net",
        "production cost vendor forced",
        "production cost final",
        "total to produce",
        "print production",
    ],
    "total_client_cost": [
        "total client net",
        "total client net includes media prod install fee",
        "total media install production net",
        "total campaign cost",
        "grand total",
        "total net cost",
    ],
    "cpm": [
        "cpm",
        "net cpm",
    ],
    "illuminated": [
        "illuminated",
        "illuminated?",
        "illuminated y/n",
        "illumination",
        "lighting",
        "lit",
    ],
    "production_forced": [
        "forced vendor production y/n",
        "forced vendor production",
        "is production forced",
        "production forced",
        "vendor forced production",
    ],
    "forced_vendor_production_cost": [
        "forced vendor production cost",
        "vendor forced production cost",
    ],
    "copy_changes": [
        "# of copy changes included at no cost after initial install",
        "number of copy changes included at no cost after initial install",
        "copy changes included",
    ],
    "paid_installs": [
        "# of paid installs",
        "number of paid installs",
        "paid installs",
    ],
    "total_postings": [
        "total postings",
        "postings",
    ],
    "print_qty": [
        "print qty per posting",
        "print quantity per posting",
        "print qty",
    ],
    "production_shipping_address": [
        "production shipping address",
        "shipping address",
        "ship to",
    ],
    "recommended_material": [
        "recommended material",
        "material",
    ],
    "creative_approval_required": [
        "creative approval required",
        "creative approval required y/n",
    ],
    "creative_due_date": [
        "creative due date",
        "creative due to vendor for approval",
        "creative asset due",
        "creative asset due to vendor",
        "creative asset due to vendor drop dead date for monday posting",
    ],
    "production_contact": [
        "production contact",
        "production contact name",
        "vendor contact name & phone #",
        "vendor contact",
    ],
    "target_area_location": [
        "target area location",
        "store covered",
        "poi",
        "target location",
    ],
    "distance_to_poi": [
        "approximate distance",
        "approximate distance mi",
        "approximate distance (mi)",
        "distance from target",
        "distance from target miles",
        "distance to poi",
        "distance to poi miles",
    ],
    "notes": [
        "notes",
        "comments",
        "comment",
        "remarks",
        "pricing comments",
        "pricing grid comments",
    ],
    "offer_id": [
        "offer id",
        "proposal id",
    ],
}


# -------------------------------------------------------------------
# Basic helpers
# -------------------------------------------------------------------

def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def _to_number(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        if isinstance(value, str):
            value = value.replace("$", "").replace(",", "").replace("%", "").strip()
            if value == "":
                return default
        return float(value)
    except Exception:
        return default


def _format_date(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%-m/%-d/%y")
    text = str(value)
    try:
        return datetime.fromisoformat(text).strftime("%-m/%-d/%y")
    except Exception:
        return text


def _campaign_date_range(requirements: dict[str, Any]) -> str:
    start = _format_date(requirements.get("campaign_start"))
    end = _format_date(requirements.get("campaign_end"))
    if start and end:
        return f"{start} - {end}"
    return ""


def _normalize_header(value: Any) -> str:
    text = _safe_str(value).lower()

    replacements = {
        "#": " number ",
        "+": " plus ",
        "&": " and ",
        "/": " ",
        "\\": " ",
        "-": " ",
        "_": " ",
        "(": " ",
        ")": " ",
        "%": " percent ",
        "?": " ",
        ".": " ",
        ",": " ",
        ":": " ",
        ";": " ",
        "\n": " ",
        "\r": " ",
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    return " ".join(text.split())


def _first_existing(row: pd.Series, candidates: list[str], default: Any = "") -> Any:
    for col in candidates:
        if col in row.index:
            val = row.get(col)
            if val is not None and not (isinstance(val, float) and pd.isna(val)):
                if str(val).strip() != "":
                    return val
    return default


def _money(value: Any) -> float:
    return round(_to_number(value, 0.0), 2)


def _int_or_blank(value: Any) -> Any:
    num = _to_number(value, None)
    if num is None:
        return ""
    return int(round(num))


def _get_unit_id(row: pd.Series) -> str:
    return _safe_str(
        _first_existing(
            row,
            [
                "unit_id",
                "unit",
                "unit_number",
                "vendor_inventory_number",
                "vendor_inventory",
                "panel_id",
            ],
        )
    )


def _is_special_unit_01134(row: pd.Series) -> bool:
    unit_id = _get_unit_id(row).replace(" ", "").upper()
    return unit_id in {"01134", "01134-SF", "1134", "1134-SF"}


def _install_cost(row: pd.Series) -> int:
    return 1600 if _is_special_unit_01134(row) else 850


def _production_cost(row: pd.Series) -> int:
    return 950 if _is_special_unit_01134(row) else 750


def _target_location_from_requirements(requirements: dict[str, Any]) -> str:
    pois = requirements.get("poi_requirements") or []
    if pois and isinstance(pois, list):
        poi = pois[0] or {}
        return _safe_str(
            poi.get("poi_name")
            or poi.get("poi_address")
            or ""
        )
    return ""


def _contracted_periods(row: pd.Series) -> Any:
    value = _first_existing(row, ["number_of_4wk_periods", "contracted_4wk_periods"], "")
    if value != "":
        return value

    contracted_weeks = _to_number(_first_existing(row, ["contracted_weeks"], 0), 0)
    if contracted_weeks:
        return round(contracted_weeks / 4, 2)

    return ""


# -------------------------------------------------------------------
# Header resolver
# -------------------------------------------------------------------

def _build_alias_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}

    for standard_field, aliases in COLUMN_ALIASES.items():
        lookup[_normalize_header(standard_field)] = standard_field
        for alias in aliases:
            lookup[_normalize_header(alias)] = standard_field

    return lookup


def _resolve_template_field(header: Any) -> str | None:
    normalized = _normalize_header(header)
    if not normalized:
        return None

    alias_lookup = _build_alias_lookup()

    if normalized in alias_lookup:
        return alias_lookup[normalized]

    for alias, standard_field in alias_lookup.items():
        if alias and (alias in normalized or normalized in alias):
            if len(alias) >= 4 and len(normalized) >= 4:
                return standard_field

    if process is not None and fuzz is not None:
        choices = list(alias_lookup.keys())
        match = process.extractOne(normalized, choices, scorer=fuzz.token_sort_ratio)

        if match:
            matched_alias, score, _ = match
            if score >= 82:
                return alias_lookup[matched_alias]

    return None


def _get_value_for_standard_field(
    field: str | None,
    header: str,
    row: pd.Series,
    requirements: dict[str, Any],
) -> Any:
    campaign_start = _format_date(requirements.get("campaign_start"))
    campaign_end = _format_date(requirements.get("campaign_end"))

    if field is None:
        h = _normalize_header(header)
        for col in row.index:
            if _normalize_header(col) == h:
                return row.get(col)
        return ""

    if field == "hold_status":
        return "Available"

    if field == "vendor":
        return "Bulletin Displays"

    if field == "availability":
        return _campaign_date_range(requirements)

    if field == "illuminated":
        return "Yes"

    if field == "production_forced":
        return "No"

    if field == "taxes":
        return 0

    if field == "market":
        return _first_existing(row, ["market", "city"], "")

    if field == "media_type":
        return _first_existing(row, ["media_type", "format"], "")

    if field == "unit_quantity":
        return 1

    if field == "unit_id":
        return _get_unit_id(row)

    if field == "geopath_id":
        return _first_existing(row, ["geopath_frame_id", "geopath_id"], "")

    if field == "location_description":
        return _first_existing(row, ["description", "location"], "")

    if field == "face":
        return _first_existing(row, ["facing", "face"], "")

    if field == "latitude":
        return _first_existing(row, ["latitude"], "")

    if field == "longitude":
        return _first_existing(row, ["longitude"], "")

    if field == "zip_code":
        return _first_existing(row, ["zip_code", "zipcode", "zip"], "")

    if field == "size":
        return _first_existing(row, ["size"], "")

    if field == "digital_spot_length":
        return _first_existing(row, ["spot_length", "spot_length_seconds"], "")

    if field == "spots_per_loop":
        return _first_existing(row, ["spots_per_loop", "number_of_spots", "spots"], "")

    if field == "loop_length":
        return _first_existing(row, ["loop_length_seconds"], "")

    if field == "digital_display_type":
        media_type = _safe_str(_first_existing(row, ["media_type"], ""))
        if "digital" in media_type.lower():
            return "Digital"
        return ""

    if field == "pixel_size":
        return _first_existing(row, ["pixel_size", "pixel_size_hxw"], "")

    if field == "weekly_impressions":
        return _int_or_blank(
            _first_existing(
                row,
                [
                    "geopath_a18_weekly_impressions",
                    "a18_weekly_impressions",
                    "weekly_impressions",
                ],
                "",
            )
        )

    if field == "four_week_impressions":
        return _int_or_blank(
            _first_existing(
                row,
                [
                    "geopath_a18_4wk_impressions",
                    "a18_4wk_impressions",
                    "contracted_impressions",
                ],
                "",
            )
        )

    if field == "total_impressions":
        return _int_or_blank(
            _first_existing(
                row,
                [
                    "contracted_impressions",
                    "geopath_a18_4wk_impressions",
                    "a18_4wk_impressions",
                ],
                "",
            )
        )

    if field == "start_date":
        return campaign_start

    if field == "end_date":
        return campaign_end

    if field == "cycle_duration":
        return "4 Week"

    if field == "number_of_cycles":
        return _contracted_periods(row)

    if field == "rate_card":
        return _money(_first_existing(row, ["rate_card_4wk", "rate_card"], 0))

    if field == "four_week_rate":
        return _money(
            _first_existing(
                row,
                [
                    "four_week_media_cost",
                    "negotiated_rate_4wk",
                    "negotiated_rate",
                ],
                0,
            )
        )

    if field == "total_media_cost":
        return _money(_first_existing(row, ["contracted_media_cost", "total_media_cost"], 0))

    if field == "install_cost":
        return _install_cost(row)

    if field == "production_cost":
        return _production_cost(row)

    if field == "total_client_cost":
        total = _to_number(_first_existing(row, ["contracted_media_cost", "total_media_cost"], 0), 0)
        total += _install_cost(row)
        total += _production_cost(row)
        return _money(total)

    if field == "cpm":
        return _money(_first_existing(row, ["cpm"], 0))

    if field == "forced_vendor_production_cost":
        return 0

    if field == "copy_changes":
        return 0

    if field == "paid_installs":
        return 1

    if field == "total_postings":
        return 1

    if field == "print_qty":
        return 1

    if field in {
        "production_shipping_address",
        "recommended_material",
        "creative_approval_required",
        "creative_due_date",
        "production_contact",
    }:
        return ""

    if field == "target_area_location":
        return _first_existing(row, ["target_location"], _target_location_from_requirements(requirements))

    if field == "distance_to_poi":
        value = _first_existing(row, ["distance_to_poi_miles"], "")
        if value == "":
            return ""
        return round(_to_number(value, 0), 2)

    if field == "notes":
        return _first_existing(
            row,
            [
                "comments",
                "comment",
                "notes",
                "pricing_grid_comments",
                "pricing_comments",
            ],
            "",
        )

    if field == "offer_id":
        return _first_existing(row, ["offer_id"], "")

    return ""


def _value_for_template_header(header: str, row: pd.Series, requirements: dict[str, Any]) -> Any:
    field = _resolve_template_field(header)
    return _get_value_for_standard_field(field, header, row, requirements)


# -------------------------------------------------------------------
# Workbook/template writing
# -------------------------------------------------------------------

def _find_header_row(ws) -> int:
    best_row = 1
    best_score = 0

    max_scan = min(ws.max_row, 30)

    for row_idx in range(1, max_scan + 1):
        fields_found = []
        raw_values = []

        for cell in ws[row_idx]:
            raw_value = _safe_str(cell.value)
            if not raw_value:
                continue

            raw_values.append(raw_value)
            field = _resolve_template_field(raw_value)

            if field:
                fields_found.append(field)

        unique_fields = set(fields_found)
        score = len(unique_fields)

        if {"vendor", "unit_id", "media_type"} & unique_fields:
            score += 2

        if {"location_description", "market"} & unique_fields:
            score += 2

        if {"four_week_rate", "rate_card", "total_media_cost"} & unique_fields:
            score += 2

        if {"install_cost", "production_cost", "total_client_cost"} & unique_fields:
            score += 2

        joined = _normalize_header(" ".join(raw_values))
        if joined in {"holds", "production"}:
            score -= 5

        if score > best_score:
            best_score = score
            best_row = row_idx

    return best_row


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col_idx in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col_idx)
        target = ws.cell(target_row, col_idx)

        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.border:
            target.border = copy(source.border)
        if source.fill:
            target.fill = copy(source.fill)
        if source.font:
            target.font = copy(source.font)


def _is_visible_row(ws, row_idx: int) -> bool:
    row_dim = ws.row_dimensions[row_idx]
    return not bool(row_dim.hidden)


def _get_visible_data_rows(ws, header_row: int, needed_rows: int) -> list[int]:
    visible_rows = []
    row_idx = header_row + 1

    max_row_to_check = max(
        ws.max_row + needed_rows + 20,
        header_row + needed_rows + 20,
    )

    while row_idx <= max_row_to_check and len(visible_rows) < needed_rows:
        if _is_visible_row(ws, row_idx):
            visible_rows.append(row_idx)
        row_idx += 1

    return visible_rows


def _clear_template_body(ws, header_row: int, rows_to_clear: int) -> None:
    visible_rows = _get_visible_data_rows(ws, header_row, rows_to_clear + 10)

    for row_idx in visible_rows:
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row_idx, col_idx).value = None


def _write_selected_to_template(ws, selected: pd.DataFrame, requirements: dict[str, Any]) -> None:
    header_row = _find_header_row(ws)

    headers = [
        ws.cell(header_row, col_idx).value
        for col_idx in range(1, ws.max_column + 1)
    ]

    needed_rows = max(len(selected), 1)
    visible_rows = _get_visible_data_rows(ws, header_row, needed_rows)

    _clear_template_body(ws, header_row, len(selected))

    if not visible_rows:
        visible_rows = [header_row + 1]

    style_source_row = visible_rows[0]

    for df_idx, (_, selected_row) in enumerate(selected.iterrows()):
        if df_idx >= len(visible_rows):
            break

        excel_row = visible_rows[df_idx]

        if excel_row != style_source_row:
            _copy_row_style(ws, style_source_row, excel_row)

        for col_idx, header in enumerate(headers, start=1):
            if header is None or str(header).strip() == "":
                continue

            value = _value_for_template_header(str(header), selected_row, requirements)
            ws.cell(excel_row, col_idx).value = value

    if selected.empty:
        ws.cell(visible_rows[0], 1).value = "No selected units."


def _write_generic_output(wb, selected: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Filled RFP Grid"

    if selected.empty:
        ws.cell(1, 1).value = "No selected units."
        return

    headers = list(selected.columns)
    ws.append(headers)

    for _, row in selected.iterrows():
        ws.append([row.get(col) for col in headers])

    ws.freeze_panes = "A2"


def write_output_workbook(
    selected: pd.DataFrame,
    excluded: pd.DataFrame,
    requirements: dict[str, Any],
    missing_fields: pd.DataFrame,
    output_path: str | Path,
    template_path: str | Path | None = None,
    column_aliases_path: str | Path | None = None,
) -> None:
    output_path = Path(output_path)

    if template_path:
        template_path = Path(template_path)
        keep_vba = template_path.suffix.lower() == ".xlsm"
        wb = load_workbook(template_path, keep_vba=keep_vba)
        ws = wb[wb.sheetnames[0]]
        _write_selected_to_template(ws, selected, requirements)
    else:
        wb = Workbook()
        _write_generic_output(wb, selected)

    wb.save(output_path)
